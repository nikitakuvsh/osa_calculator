from copy import copy

import pandas as pd
from openpyxl import load_workbook


def find_dmr_header(source_file):
    """
    Автоматически ищет строку заголовков на листе DMR.
    """

    raw = pd.read_excel(
        source_file,
        sheet_name="DMR",
        header=None
    )

    for i in range(len(raw)):

        row = [
            str(x).strip()
            for x in raw.iloc[i].tolist()
        ]

        if (
            "Client" in row
            and "p1" in row
            and "p13" in row
        ):
            return i

    raise Exception(
        "Не удалось найти заголовки на листе DMR"
    )

def find_export_header(source_file):
    """
    Автоматически ищет строку заголовков на первом листе.
    """

    raw = pd.read_excel(
        source_file,
        sheet_name=0,
        header=None
    )

    required = {
        "full_period_name",
        "OSA After",
        "OSA Before",
        "Merch Impact",
        "retailer_name"
    }

    for i in range(len(raw)):

        row = {
            str(x).strip()
            for x in raw.iloc[i].tolist()
            if pd.notna(x)
        }

        if required.issubset(row):
            return i

    raise Exception(
        "Не удалось найти строку заголовков на листе Export"
    )


def merge_data(workbook, source_file):


    export_df = pd.read_excel(
        source_file,
        header=find_export_header(source_file)
    )

    print("    Найдены заголовки Export")

    # -----------------------------
    # DMR
    # -----------------------------

    dmr_header = find_dmr_header(source_file)

    dmr_df = pd.read_excel(
        source_file,
        sheet_name="DMR",
        header=dmr_header
    )

    dmr_df.columns = (
        dmr_df.columns
        .astype(str)
        .str.strip()
    )

    print(f"    Найдены заголовки DMR")

    # -----------------------------
    # Копируем первый лист в Export
    # -----------------------------

    source_wb = load_workbook(source_file)

    source_ws = source_wb.worksheets[0]

    if "Export" not in workbook.sheetnames:
        ws = workbook.create_sheet("Export")
    else:
        ws = workbook["Export"]

    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    for row in source_ws.iter_rows():

        for cell in row:

            target = ws[cell.coordinate]

            target.value = cell.value

            if cell.has_style:
                target._style = copy(cell._style)

            target.number_format = cell.number_format

            if cell.alignment:
                target.alignment = copy(cell.alignment)

            if cell.border:
                target.border = copy(cell.border)

            if cell.fill:
                target.fill = copy(cell.fill)

            if cell.font:
                target.font = copy(cell.font)

    for col, dim in source_ws.column_dimensions.items():
        ws.column_dimensions[col].width = dim.width

    for row, dim in source_ws.row_dimensions.items():
        ws.row_dimensions[row].height = dim.height

    print(
        f"    Export заполнен: {source_ws.max_row} строк, "
        f"    {source_ws.max_column} колонок"
    )

    dmr_dict = {}
    period_columns = [f"p{i}" for i in range(1,14)]
    for _, row in dmr_df.iterrows():
        retailer = str(row["Client"]).strip()
        dmr_dict[retailer] = {}
        for period in period_columns:
            value = row.get(period)
            if pd.notna(value):
                dmr_dict[retailer][period] = float(value)

    print(
        f"    DMR найден. Строк: {len(dmr_df)}"
    )

    return export_df, dmr_dict