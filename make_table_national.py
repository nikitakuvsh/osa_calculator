from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import re
import pandas as pd
from forecast import (
    forecast_missing_periods,
    forecast_missing_periods_osa,
    forecast_next_year,
    forecast_next_year_osa
)
from time import sleep

logs = {
    "merge": False,
    "forecast": False,
    "dif_merch": False,
    "dif_osa": False,
    "lsv_merch": False,
    "lsv_osa": False,
    "total_osa": False
}


def create_table(wb, df, dmr_dict):

    if "Report" not in wb.sheetnames:
        ws = wb.create_sheet("Report")
    else:
        ws = wb["Report"]

        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
                cell._style = None

        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged))


    # ======================================================
    # Подготовка данных
    # ======================================================

    df = df[
        df["retailer_name"].notna()
        &
        (df["retailer_name"].astype(str).str.strip() != "")
    ]


    retailers = df["retailer_name"].drop_duplicates().tolist()

    # Фильр только с DMR
    retailers = [retailer for retailer in retailers if retailer in dmr_dict and any(isinstance(v, (int, float)) for v in dmr_dict[retailer].values())]
    if not retailers: return wb
    print(f'    Количество retailers после фильтра по DMR = {len(retailers)}')


    if not retailers:
        return wb

    raw_periods = sorted(df["full_period_name"].dropna().unique().tolist())


    # год из файла
    match = re.search(
        r"(20\d{2})",
        str(raw_periods[0])
    )

    current_year = int(match.group(1)) if match else 2026

    periods = [f"{current_year} P{i:02d}" for i in range(1, 14)]



    # ======================================================
    # Начало таблицы
    # ======================================================

    START_ROW = 3
    START_COL = 3



    # ======================================================
    # Размеры
    # ======================================================

    ws.column_dimensions[
        get_column_letter(START_COL)
    ].width = 28


    ws.column_dimensions[
        get_column_letter(START_COL + 1)
    ].width = 28


    for i in range(13):
        ws.column_dimensions[
            get_column_letter(START_COL + 2 + i)
        ].width = 12


    ws.column_dimensions[
        get_column_letter(START_COL + 15)
    ].width = 14



        # ======================================================
    # Стили
    # ======================================================

    green = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3"   # светло-зелёный
    )

    retailer_fill = PatternFill(
        fill_type="solid",
        fgColor="B4C7E7"   # спокойный голубой
    )

    white = PatternFill(
        fill_type="solid",
        fgColor="FFFFFF"
    )


    thin = Side(
        style="thin",
        color="000000"
    )

    thick = Side(
        style="medium",
        color="7F7F7F"
    )


    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )


    center = Alignment(
        horizontal="center",
        vertical="center"
    )


    bold = Font(
        bold=True,
        color="1F1F1F"
    )


    bold_red = Font(
        bold=True,
        color="FF0000"
    )


    bold_white = Font(
        bold=True,
        color="FFFFFF"
    )

    red_text = Font(
        bold=True,
        color="FF0000"
    )

    # ======================================================
    # Создание блоков retailer
    # ======================================================


    current_row = START_ROW


    for retailer in retailers:
        dmr_values = dmr_dict.get(retailer, {})
        if not dmr_values: continue


        # ------------------------------
        # Заголовок retailer
        # ------------------------------

        for i, text in enumerate(
            [f"P{i}" for i in range(1,14)] + ["total"]
        ):

            cell = ws.cell(
                row=current_row,
                column=START_COL + 2 + i,
                value=text
            )

            cell.font = bold
            cell.alignment = center


        cell = ws.cell(
            row=current_row + 1,
            column=START_COL,
            value=retailer
        )

        cell.fill = retailer_fill
        cell.font = bold_white
        cell.alignment = center



        # ------------------------------
        # Левая группа
        # ------------------------------

        ws.merge_cells(
            start_row=current_row + 2,
            start_column=START_COL,
            end_row=current_row + 7,
            end_column=START_COL
        )


        cell = ws.cell(
            current_row + 2,
            START_COL
        )

        cell.fill = green
        cell.alignment = center



        # ------------------------------
        # OSA before
        # ------------------------------

        ws.merge_cells(
            start_row=current_row + 8,
            start_column=START_COL,
            end_row=current_row + 12,
            end_column=START_COL
        )


        cell = ws.cell(
            current_row + 8,
            START_COL,
            value="OSA before"
        )

        cell.font = bold
        cell.alignment = center
        cell.fill = green



        # ------------------------------
        # OSA
        # ------------------------------

        ws.merge_cells(
            start_row=current_row + 13,
            start_column=START_COL,
            end_row=current_row + 16,
            end_column=START_COL
        )


        cell = ws.cell(
            current_row + 13,
            START_COL,
            value="OSA"
        )

        cell.fill = green
        cell.font = bold
        cell.alignment = center



        # ------------------------------
        # Названия строк
        # ------------------------------

        labels = [

            "DMR",
            "",
            f"Merch Impact {current_year}",
            f"Merch Impact {current_year + 1}",
            "dif",
            "ЭФ-т LSV, %",
            "ЭФ-т, млн руб Merch in",

            f"osa before {current_year}",
            f"osa before {current_year + 1}",
            "dif",
            "ЭФ-т LSV, %",
            "ЭФ-т, млн руб Osa Base",

            f"total OSA {current_year}",
            f"total OSA {current_year + 1}",
            "dif",
            f"total {current_year + 1} млн руб"
        ]



        green_rows = {
            6,     # ЭФ-т, млн руб Merch in
            11,    # ЭФ-т, млн руб Osa Base
            15     # total 2026 млн руб
        }


        for i,label in enumerate(labels):

            cell = ws.cell(
                row=current_row + 1 + i,
                column=START_COL + 1,
                value=label
            )


            if i in green_rows:
                cell.fill = green
                cell.font = bold

            else:
                cell.fill = white


            if i in [0, 12, 13]:
                cell.font = bold

            if i in [17]: cell.fill = white


            cell.alignment = center


            cell.alignment = center

        # ==================================================
        # Красивое форматирование строк
        # ==================================================

        full_rows = [
            6,   # ЭФ-т, млн руб Merch in
            11,  # ЭФ-т, млн руб Osa Base
            15   # total 2027 млн руб
        ]


        bold_rows = [
            0,   # DMR
            12,  # total OSA 2026
            13   # total OSA 2027
        ]


        for row_index in full_rows:

            excel_row = current_row + 1 + row_index

            for col in range(
                START_COL,
                START_COL + 16
            ):

                cell = ws.cell(
                    row=excel_row,
                    column=col
                )

                cell.fill = green
                
                cell.font = bold



        for row_index in bold_rows:

            excel_row = current_row + 1 + row_index

            for col in range(
                START_COL,
                START_COL + 16
            ):

                cell = ws.cell(
                    row=excel_row,
                    column=col
                )

                cell.font = bold



        # ==================================================
        # Заполнение данных
        # ==================================================

        merch_values = []
        osa_values = []

        if not logs["merge"]:
            print('    Мержу данные со вкладкой Export')
            logs["merge"] = True
            sleep(0.5)        

        for period_index, period in enumerate(periods):

            col = START_COL + 2 + period_index
            period_name = f"p{period_index + 1}"
            dmr = dmr_values.get(period_name)
            if dmr is not None:
                ws.cell(
                    row=current_row + 1,
                    column=col,
                    value=dmr
                ).number_format = '# ##0'
                


            row_osa = current_row + 7
            row_merch = current_row + 4



            data = df[
                (df["retailer_name"] == retailer)
                &
                (df["full_period_name"] == period)
            ]
        



            if not data.empty:


                osa = data["OSA Before"].iloc[0]


                if pd.notna(osa):
                    osa = float(osa)
                    osa_values.append(osa)

                    ws.cell(
                        row=row_osa + 1,
                        column=col,
                        value=round(float(osa),4)
                    ).number_format = "0.00%"

                else: osa_values.append(None)



                merch = data["Merch Impact"].iloc[0]


                if pd.notna(merch):
                    merch = float(merch)
                    merch_values.append(merch)

                    ws.cell(
                        row=row_merch - 1,
                        column=col,
                        value=round(float(merch),4)
                    ).number_format = "0.00%"

                else: merch_values.append(None)

            values = [v for v in dmr_values.values() if isinstance(v, (int, float))]
            if values:
                ws.cell(
                    row=current_row + 1,
                    column=START_COL + 15,
                    value=sum(values)
                ).number_format = "# ##0"

        # ==================================================
        # Заполняем отсутствующие периоды текущего года
        # ==================================================

        while len(merch_values) < 13: merch_values.append(None)
        while len(osa_values) < 13: osa_values.append(None)

        merch_2026_full = forecast_missing_periods(merch_values)
        osa_2026_full = forecast_missing_periods_osa(osa_values)

        for i in range(13):
            col = START_COL + 2 + i

            cell = ws.cell(
                row=current_row + 3,
                column=col
            )

            if cell.value is None:
                cell.value = merch_2026_full[i]
                cell.number_format = "0.00%"
                cell.font = red_text


        for i in range(13):
            col = START_COL + 2 + i

            cell = ws.cell(
                row=current_row + 8,
                column=col
            )

            if cell.value is None:
                cell.value = osa_2026_full[i]
                cell.number_format = "0.00%"
                cell.font = red_text

        for period_index in range(13):
            col = START_COL + 2 + period_index
            period_name = f"p{period_index + 1}"
            dmr = dmr_values.get(period_name)
            if dmr is not None:
                ws.cell(
                    row=current_row + 1,
                    column=col,
                    value=dmr
                ).number_format = "# ##0"

        if not logs["forecast"]:
            print('    Занимаюсь творческим выдумыванием процентов на следующий год :)')
            logs["forecast"] = True
            sleep(0.5)

        # Берём уже заполненный полностью текущий год
        full_merch_values = []

        for i in range(13):

            value = ws.cell(
                row=current_row + 3,
                column=START_COL + 2 + i
            ).value

            if value is not None:
                full_merch_values.append(value)



        full_osa_values = []

        for i in range(13):

            value = ws.cell(
                row=current_row + 8,
                column=START_COL + 2 + i
            ).value

            if value is not None:
                full_osa_values.append(value)



        # прогноз следующего года
        merch_forecast = forecast_next_year(full_merch_values)
        osa_forecast = forecast_next_year_osa(full_osa_values)

        for i in range(13):
            value = merch_forecast[i] if i < len(merch_forecast) else None

            if value is not None:
                cell = ws.cell(
                    row=current_row + 4,
                    column=START_COL + 2 + i,
                    value=value
                )

                cell.number_format = "0.00%"
                cell.font = red_text

        for i in range(13):
            value = osa_forecast[i] if i < len(osa_forecast) else None

            if value is not None:
                cell = ws.cell(
                    row=current_row + 9,
                    column=START_COL + 2 + i,
                    value=value
                )

                cell.number_format = "0.00%"
                cell.font = red_text

        # ==================================================
        # DIF Merch Impact
        # ==================================================

        if not logs["dif_merch"]:
            print('    Считаю разницу dif Merch Impact')
            logs["dif_merch"] = True
            sleep(0.7)

        for i in range(13):

            col = START_COL + 2 + i

            current_merch = ws.cell(
                row=current_row + 3,
                column=col
            ).value


            forecast_merch = ws.cell(
                row=current_row + 4,
                column=col
            ).value


            if current_merch is not None and forecast_merch is not None:

                ws.cell(
                    row=current_row + 5,
                    column=col,
                    value=round(
                        forecast_merch - current_merch,
                        4
                    )
                ).number_format = "0.00%"



        # ==================================================
        # DIF OSA
        # ==================================================

        if not logs["dif_osa"]:
            print('    Считаю разницу dif OSA')
            logs["dif_osa"] = True
            sleep(0.3)

        for i in range(13):

            col = START_COL + 2 + i

            current_osa = ws.cell(
                row=current_row + 8,
                column=col
            ).value


            forecast_osa = ws.cell(
                row=current_row + 9,
                column=col
            ).value


            if current_osa is not None and forecast_osa is not None:

                ws.cell(
                    row=current_row + 10,
                    column=col,
                    value=round(
                        forecast_osa - current_osa,
                        4
                    )
                ).number_format = "0.00%"

        # ==================================================
        # ЭФ-т LSV Merch Impact
        # ==================================================

        if not logs["lsv_merch"]:
            print('    Рассчитываю Эф-т LSV Merch Impact')
            logs["lsv_merch"] = True
            sleep(1)

        for i in range(13):

            col = START_COL + 2 + i

            dif = ws.cell(
                row=current_row + 5,
                column=col
            ).value


            if dif is not None:

                ws.cell(
                    row=current_row + 6,
                    column=col,
                    value=round(
                        dif / 0.03 * 0.01,
                        4
                    )
                ).number_format = "0.00%"

        # ==================================================
        # ЭФ-т, млн руб Merch in
        # ==================================================

        for i in range(13):

            col = START_COL + 2 + i

            dmr = ws.cell(
                row=current_row + 1,
                column=col
            ).value

            lsv = ws.cell(
                row=current_row + 6,
                column=col
            ).value

            if dmr is not None and lsv is not None:

                ws.cell(
                    row=current_row + 7,
                    column=col,
                    value=round(dmr * lsv, 2)
                ).number_format = '# ##0.00'



        # ==================================================
        # ЭФ-т LSV OSA
        # ==================================================

        
        if not logs["lsv_osa"]:
            print('    Рассчитываю Эф-т LSV OSA')
            logs["lsv_osa"] = True
            sleep(0.4)

        for i in range(13):

            col = START_COL + 2 + i

            dif = ws.cell(
                row=current_row + 10,
                column=col
            ).value


            if dif is not None:

                ws.cell(
                    row=current_row + 11,
                    column=col,
                    value=round(
                        dif / 0.03 * 0.01,
                        4
                    )
                ).number_format = "0.00%"

        # ==================================================
        # ЭФ-т, млн руб Osa Base
        # ==================================================

        for i in range(13):

            col = START_COL + 2 + i

            dmr = ws.cell(
                row=current_row + 1,
                column=col
            ).value

            lsv = ws.cell(
                row=current_row + 11,
                column=col
            ).value

            if dmr is not None and lsv is not None:

                ws.cell(
                    row=current_row + 12,
                    column=col,
                    value=round(dmr * lsv, 2)
                ).number_format = '# ##0.00'

        # ==================================================
        # TOTAL OSA
        # ==================================================
                
        if not logs["total_osa"]:
            print('    Считаю Total OSA')
            logs["total_osa"] = True
            sleep(1)

        # ==================================================
        # TOTAL OSA
        # ==================================================

        for i in range(13):

            col = START_COL + 2 + i


            merch_current = ws.cell(
                row=current_row + 3,
                column=col
            ).value


            osa_current = ws.cell(
                row=current_row + 8,
                column=col
            ).value


            if merch_current is not None and osa_current is not None:

                total_current = min(
                    merch_current + osa_current,
                    1
                )

                ws.cell(
                    row=current_row + 13,
                    column=col,
                    value=round(total_current,4)
                ).number_format = "0.00%"



            merch_next = ws.cell(
                row=current_row + 4,
                column=col
            ).value


            osa_next = ws.cell(
                row=current_row + 9,
                column=col
            ).value


            if merch_next is not None and osa_next is not None:

                total_next = min(
                    merch_next + osa_next,
                    1
                )

                ws.cell(
                    row=current_row + 14,
                    column=col,
                    value=round(total_next,4)
                ).number_format = "0.00%"



            total_current_value = ws.cell(
                row=current_row + 13,
                column=col
            ).value


            total_next_value = ws.cell(
                row=current_row + 14,
                column=col
            ).value



            if total_current_value is not None and total_next_value is not None:

                ws.cell(
                    row=current_row + 15,
                    column=col,
                    value=round(
                        total_next_value - total_current_value,
                        4
                    )
                ).number_format = "0.00%"
            

        # красим аналитические значения
        for r in [
            current_row + 4,   # Merch Impact следующий год
            current_row + 9,   # OSA следующий год
        ]:

            for c in range(
                START_COL + 2,
                START_COL + 16
            ):
                ws.cell(
                    row=r,
                    column=c
                ).font = bold_red

        for r in [
            current_row + 7,
            current_row + 12,
            current_row + 16
        ]:

            for c in range(
                START_COL,
                START_COL + 16
            ):

                ws.cell(
                    row=r,
                    column=c
                ).fill = green

                ws.cell(
                    row=r,
                    column=c
                ).font = bold

        # ==================================================
        # TOTAL COLUMN
        # ==================================================

        total_col = START_COL + 15


        # ------------------------------
        # Merch Impact current_year
        # среднее всех месяцев
        # ------------------------------

        merch_current_values = []

        for i in range(13):

            value = ws.cell(
                row=current_row + 3,
                column=START_COL + 2 + i
            ).value

            if value is not None:
                merch_current_values.append(value)


        if merch_current_values:

            ws.cell(
                row=current_row + 3,
                column=total_col,
                value=round(
                    sum(merch_current_values) / len(merch_current_values),
                    4
                )
            ).number_format = "0.00%"



        # ------------------------------
        # Merch Impact next_year
        # ------------------------------

        merch_next_values = []

        for i in range(13):

            value = ws.cell(
                row=current_row + 4,
                column=START_COL + 2 + i
            ).value

            if value is not None:
                merch_next_values.append(value)


        if merch_next_values:

            ws.cell(
                row=current_row + 4,
                column=total_col,
                value=round(
                    sum(merch_next_values) / len(merch_next_values),
                    4
                )
            ).number_format = "0.00%"



        # ------------------------------
        # DIF Merch
        # ------------------------------

        merch_current_total = ws.cell(
            row=current_row + 3,
            column=total_col
        ).value


        merch_next_total = ws.cell(
            row=current_row + 4,
            column=total_col
        ).value


        if merch_current_total is not None and merch_next_total is not None:

            ws.cell(
                row=current_row + 5,
                column=total_col,
                value=round(
                    merch_next_total - merch_current_total,
                    4
                )
            ).number_format = "0.00%"



        # ------------------------------
        # LSV Merch
        # ------------------------------

        dif_merch_total = ws.cell(
            row=current_row + 5,
            column=total_col
        ).value


        if dif_merch_total is not None:

            ws.cell(
                row=current_row + 6,
                column=total_col,
                value=round(
                    dif_merch_total / 0.03 * 0.01,
                    4
                )
            ).number_format = "0.00%"





        # ==================================================
        # OSA TOTAL COLUMN
        # ==================================================


        osa_current_values = []

        for i in range(13):

            value = ws.cell(
                row=current_row + 8,
                column=START_COL + 2 + i
            ).value

            if value is not None:
                osa_current_values.append(value)



        if osa_current_values:

            ws.cell(
                row=current_row + 8,
                column=total_col,
                value=round(
                    sum(osa_current_values) / len(osa_current_values),
                    4
                )
            ).number_format = "0.00%"




        osa_next_values = []

        for i in range(13):

            value = ws.cell(
                row=current_row + 9,
                column=START_COL + 2 + i
            ).value

            if value is not None:
                osa_next_values.append(value)



        if osa_next_values:

            ws.cell(
                row=current_row + 9,
                column=total_col,
                value=round(
                    sum(osa_next_values) / len(osa_next_values),
                    4
                )
            ).number_format = "0.00%"




        osa_current_total = ws.cell(
            row=current_row + 8,
            column=total_col
        ).value


        osa_next_total = ws.cell(
            row=current_row + 9,
            column=total_col
        ).value


        # DIF OSA

        if osa_current_total is not None and osa_next_total is not None:

            ws.cell(
                row=current_row + 10,
                column=total_col,
                value=round(
                    osa_next_total - osa_current_total,
                    4
                )
            ).number_format = "0.00%"



        # LSV OSA

        dif_osa_total = ws.cell(
            row=current_row + 10,
            column=total_col
        ).value


        if dif_osa_total is not None:

            ws.cell(
                row=current_row + 11,
            column=total_col,
                value=round(
                    dif_osa_total / 0.03 * 0.01,
                    4
                )
            ).number_format = "0.00%"





        # ==================================================
        # TOTAL OSA
        # ==================================================

        total_osa_current_values = []
        total_osa_next_values = []


        for i in range(13):

            current = ws.cell(
                row=current_row + 13,
                column=START_COL + 2 + i
            ).value

            next_year = ws.cell(
                row=current_row + 14,
                column=START_COL + 2 + i
            ).value


            if current is not None:
                total_osa_current_values.append(current)

            if next_year is not None:
                total_osa_next_values.append(next_year)



        if total_osa_current_values:

            ws.cell(
                row=current_row + 13,
                column=total_col,
                value=round(
                    sum(total_osa_current_values) /
                    len(total_osa_current_values),
                    4
                )
            ).number_format = "0.00%"



        if total_osa_next_values:

            ws.cell(
                row=current_row + 14,
                column=total_col,
                value=round(
                    sum(total_osa_next_values) /
                    len(total_osa_next_values),
                    4
                )
            ).number_format = "0.00%"



        # DIF TOTAL OSA

        total_current = ws.cell(
            row=current_row + 13,
            column=total_col
        ).value


        total_next = ws.cell(
            row=current_row + 14,
            column=total_col
        ).value


        if total_current is not None and total_next is not None:

            ws.cell(
                row=current_row + 15,
                column=total_col,
                value=round(
                    total_next - total_current,
                    4
                )
            ).number_format = "0.00%"

        # ==================================================
        # total 2027 млн руб
        # ==================================================

        for i in range(13):

            col = START_COL + 2 + i

            merch_effect = ws.cell(
                row=current_row + 7,
                column=col
            ).value

            osa_effect = ws.cell(
                row=current_row + 12,
                column=col
            ).value

            if merch_effect is not None and osa_effect is not None:

                ws.cell(
                    row=current_row + 16,
                    column=col,
                    value=round(
                        merch_effect + osa_effect,
                        2
                    )
                ).number_format = "# ##0.00"

        # Эф-т, млн руб Merch in (total)

        # ==================================================
        # TOTAL ЭФ-Т млн руб Merch in
        # сумма всех периодов
        # ==================================================

        merch_effect_total = []

        for i in range(13):

            value = ws.cell(
                row=current_row + 7,
                column=START_COL + 2 + i
            ).value

            if value is not None:
                merch_effect_total.append(value)


        if merch_effect_total:

            ws.cell(
                row=current_row + 7,
                column=total_col,
                value=round(
                    sum(merch_effect_total),
                    2
                )
            ).number_format = "# ##0.00"



        # ==================================================
        # TOTAL ЭФ-Т млн руб Osa Base
        # сумма всех периодов
        # ==================================================

        osa_effect_total = []

        for i in range(13):

            value = ws.cell(
                row=current_row + 12,
                column=START_COL + 2 + i
            ).value

            if value is not None:
                osa_effect_total.append(value)


        if osa_effect_total:

            ws.cell(
                row=current_row + 12,
                column=total_col,
                value=round(
                    sum(osa_effect_total),
                    2
                )
            ).number_format = "# ##0.00"



        # ==================================================
        # TOTAL current_year + 1 млн руб
        # Merch + OSA
        # ==================================================

        merch_total = ws.cell(
            row=current_row + 7,
            column=total_col
        ).value


        osa_total = ws.cell(
            row=current_row + 12,
            column=total_col
        ).value


        if merch_total is not None and osa_total is not None:

            ws.cell(
                row=current_row + 16,
                column=total_col,
                value=round(
                    merch_total + osa_total,
                    2
                )
            ).number_format = "# ##0.00"

# ==================================================
# Бордеры блока
# ==================================================

        for r in range(
            current_row,
            current_row + 17
        ):

            for c in range(
                START_COL,
                START_COL + 16
            ):

                cell = ws.cell(
                    row=r,
                    column=c
                )

                # обычная сетка
                cell.border = border
                cell.alignment = center


                # жирные разделители
                if r in [
                    current_row + 7,   # ЭФ-т, млн руб Merch in
                    current_row + 12,  # ЭФ-т, млн руб Osa Base / перед total OSA
                    current_row + 16   # dif после total 2026
                ]:

                    cell.border = Border(
                        left=thin,
                        right=thin,
                        top=thin,
                        bottom=thick
                    )



        # следующий retailer ниже
        current_row += 19


    print('    Ввожу последние штрихи')
    sleep(1)
    return wb