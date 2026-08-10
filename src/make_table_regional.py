from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import re
import pandas as pd
from forecast import (
    forecast_missing_periods,
    forecast_missing_periods_osa,
    forecast_next_year,
    forecast_next_year_osa
)
from time import sleep

logs = {
    "merge": False,
    "forecast": False,
    "dif_merch": False,
    "dif_osa": False,
    "lsv_merch": False,
    "lsv_osa": False,
    "total_osa": False
}


def create_table(wb, df, dmr_dict):
    """
    Создаёт компактную аналитическую таблицу по demand_region.

    Структура:

        demand_region
            retailer
                DMR
                OSA After current_year
                OSA After current_year + 1
                dif
                ЭФ-т LSV, %
                ЭФ-т, млн руб

    ВАЖНО:
    - OSA After current_year берётся из исходного DataFrame.
    - OSA After current_year + 1 рассчитывается Python.
    - dif, LSV и ЭФ-т рассчитываются Excel-формулами.
    - Одинаковая компания в разных demand_region считается отдельно.
    - Пустые строки/компании без DMR не выводятся.
    """

    # ==========================================================
    # Подготовка листа
    # ==========================================================

    if "Report" not in wb.sheetnames:
        ws = wb.create_sheet("Report")
    else:
        ws = wb["Report"]

        # Очищаем старый лист
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
                cell._style = copy.copy(ws["A1"]._style)

        # Удаляем старые объединения
        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged))

    # ==========================================================
    # Проверяем колонки
    # ==========================================================

    required_columns = [
        "retailer_name",
        "demand_region",
        "full_period_name",
        "OSA After"
    ]

    missing_columns = [
        column
        for column in required_columns
        if column not in df.columns
    ]

    if missing_columns:
        raise KeyError(
            f"В DataFrame отсутствуют колонки: {missing_columns}"
        )

    # ==========================================================
    # Чистим исходные данные
    # ==========================================================

    df = df[
        df["retailer_name"].notna()
        &
        (df["retailer_name"].astype(str).str.strip() != "")
        &
        df["demand_region"].notna()
        &
        (df["demand_region"].astype(str).str.strip() != "")
    ].copy()

    if df.empty:
        return wb

    df["retailer_name"] = (
        df["retailer_name"]
        .astype(str)
        .str.strip()
    )

    df["demand_region"] = (
        df["demand_region"]
        .astype(str)
        .str.strip()
    )

    # ==========================================================
    # Определяем текущий год
    # ==========================================================

    raw_periods = (
        df["full_period_name"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    if not raw_periods:
        return wb

    match = re.search(r"(20\d{2})", raw_periods[0])

    current_year = (
        int(match.group(1))
        if match
        else 2026
    )

    next_year = current_year + 1

    periods = [
        f"{current_year} P{i:02d}"
        for i in range(1, 14)
    ]

    # ==========================================================
    # Параметры таблицы
    # ==========================================================

    START_ROW = 3
    START_COL = 3

    # C = retailer
    # D = показатель
    # E:Q = P1:P13
    # R = total

    TOTAL_COL = START_COL + 15

    # ==========================================================
    # Ширина колонок
    # ==========================================================

    ws.column_dimensions[
        get_column_letter(START_COL)
    ].width = 28

    ws.column_dimensions[
        get_column_letter(START_COL + 1)
    ].width = 28

    for i in range(13):
        ws.column_dimensions[
            get_column_letter(
                START_COL + 2 + i
            )
        ].width = 17

    ws.column_dimensions[
        get_column_letter(TOTAL_COL)
    ].width = 17

    # ==========================================================
    # Стили
    # ==========================================================

    green = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3"
    )

    dark_green = PatternFill(
        fill_type="solid",
        fgColor="A9D18E"
    )

    retailer_fill = PatternFill(
        fill_type="solid",
        fgColor="B4C7E7"
    )

    white = PatternFill(
        fill_type="solid",
        fgColor="FFFFFF"
    )

    thin = Side(
        style="thin",
        color="000000"
    )

    medium = Side(
        style="medium",
        color="7F7F7F"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    left = Alignment(
        horizontal="left",
        vertical="center"
    )

    bold = Font(
        bold=True,
        color="1F1F1F"
    )

    red_text = Font(
        bold=True,
        color="FF0000"
    )

    # ==========================================================
    # Получение DMR
    # ==========================================================

    def get_dmr_values(region, retailer):
        """
        Поддерживает:

        1.
        dmr_dict[(region, retailer)]

        2.
        dmr_dict[region][retailer]

        3.
        dmr_dict[retailer]
        """

        # ----------------------------------------------
        # (region, retailer)
        # ----------------------------------------------

        key = (region, retailer)

        if key in dmr_dict:
            value = dmr_dict[key]

            if isinstance(value, dict):
                return value

        # ----------------------------------------------
        # region -> retailer
        # ----------------------------------------------

        region_data = dmr_dict.get(region)

        if isinstance(region_data, dict):

            retailer_data = region_data.get(retailer)

            if isinstance(retailer_data, dict):
                return retailer_data

        # ----------------------------------------------
        # retailer
        # ----------------------------------------------

        retailer_data = dmr_dict.get(retailer)

        if isinstance(retailer_data, dict):
            return retailer_data

        return {}

    # ==========================================================
    # Уникальные пары region + retailer
    # ==========================================================

    grouped = (
        df[
            [
                "demand_region",
                "retailer_name"
            ]
        ]
        .drop_duplicates()
    )

    if grouped.empty:
        return wb

    regions = (
        grouped["demand_region"]
        .drop_duplicates()
        .tolist()
    )

    current_row = START_ROW

    # ==========================================================
    # Регионы
    # ==========================================================

    for region in regions:

        region_data = grouped[
            grouped["demand_region"] == region
        ]

        retailers = (
            region_data["retailer_name"]
            .drop_duplicates()
            .tolist()
        )

        # ------------------------------------------------------
        # Оставляем только компании, у которых есть DMR
        # ------------------------------------------------------

        valid_retailers = []

        for retailer in retailers:

            dmr_values = get_dmr_values(
                region,
                retailer
            )

            has_dmr = any(
                isinstance(value, (int, float))
                and not isinstance(value, bool)
                for value in dmr_values.values()
            )

            if has_dmr:
                valid_retailers.append(retailer)

        if not valid_retailers:
            continue

        # ======================================================
        # ЗЕЛЁНАЯ ПОЛОСА REGION + ИТОГИ ПО РЕГИОНУ
        # ======================================================

        region_row = current_row

        # ------------------------------------------------------
        # Получаем все компании региона
        # ------------------------------------------------------

        region_retailers = (
            region_data["retailer_name"]
            .drop_duplicates()
            .tolist()
        )

        # ------------------------------------------------------
        # Собираем DMR по всем компаниям региона
        # ------------------------------------------------------

        region_dmr_totals = {
            f"p{i}": 0
            for i in range(1, 14)
        }

        for retailer in region_retailers:

            retailer_dmr = get_dmr_values(
                region,
                retailer
            )

            for i in range(1, 14):

                period_key = f"p{i}"
                value = retailer_dmr.get(period_key)

                if (
                    value is not None
                    and isinstance(value, (int, float))
                    and not isinstance(value, bool)
                ):
                    region_dmr_totals[period_key] += float(value)

        # ------------------------------------------------------
        # Название региона
        # ------------------------------------------------------

        region_cell = ws.cell(
            row=region_row,
            column=START_COL,
            value=region
        )

        region_cell.fill = dark_green
        region_cell.font = bold
        region_cell.alignment = left

        # ------------------------------------------------------
        # D оставляем пустым
        # ------------------------------------------------------

        region_label_cell = ws.cell(
            row=region_row,
            column=START_COL + 1
        )

        region_label_cell.fill = dark_green

        # ------------------------------------------------------
        # Значения DMR по периодам
        #
        # ВАЖНО:
        # здесь НЕТ P1/P2/P3...
        # только сами значения
        # ------------------------------------------------------

        for i in range(1, 14):

            col = START_COL + 1 + i

            value = region_dmr_totals[f"p{i}"]

            cell = ws.cell(
                row=region_row,
                column=col,
                value=value if value != 0 else None
            )

            cell.fill = dark_green
            cell.font = bold
            cell.number_format = "# ##0.00"
            cell.alignment = center

        for col in range(1, 150):
            ws.cell(row=region_row, column=col).fill = dark_green

        # ------------------------------------------------------
        # TOTAL региона
        # ------------------------------------------------------

        region_total = sum(
            region_dmr_totals.values()
        )

        total_cell = ws.cell(
            row=region_row,
            column=TOTAL_COL,
            value=region_total if region_total != 0 else None
        )

        total_cell.fill = dark_green
        total_cell.font = bold
        total_cell.number_format = "# ##0.00"
        total_cell.alignment = center

        # ------------------------------------------------------
        # ЗЕЛЁНАЯ ПОЛОСКА ОТ C ДО R
        # ------------------------------------------------------

        for col in range(
            START_COL,
            TOTAL_COL + 1
        ):

            cell = ws.cell(
                row=region_row,
                column=col
            )

            cell.fill = dark_green
            cell.border = Border(
                left=thin,
                right=thin,
                top=medium,
                bottom=medium
            )

        # ------------------------------------------------------
        # Переходим к компаниям
        # ------------------------------------------------------

        current_row += 1

        # ======================================================
        # Компании
        # ======================================================

        for retailer in valid_retailers:

            dmr_values = get_dmr_values(
                region,
                retailer
            )

            header_row = current_row

            # --------------------------------------------------
            # Заголовки периодов
            # --------------------------------------------------

            headers = [
                f"P{i}"
                for i in range(1, 14)
            ] + ["total"]

            for i, header in enumerate(headers):

                cell = ws.cell(
                    row=header_row,
                    column=START_COL + 2 + i,
                    value=header
                )

                cell.font = bold
                cell.alignment = center
                cell.fill = white
                cell.border = border

            # --------------------------------------------------
            # Компания
            # --------------------------------------------------

            # ==========================================================
            # Название компании — объединяем на 6 строк
            # ==========================================================

            retailer_start_row = header_row + 1
            retailer_end_row = header_row + 6

            ws.merge_cells(
                start_row=retailer_start_row,
                start_column=START_COL,
                end_row=retailer_end_row,
                end_column=START_COL
            )

            retailer_cell = ws.cell(
                row=retailer_start_row,
                column=START_COL,
                value=retailer
            )

            retailer_cell.fill = retailer_fill
            retailer_cell.font = bold
            retailer_cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            for row in range(
                retailer_start_row,
                retailer_end_row + 1
            ):
                cell = ws.cell(
                    row=row,
                    column=START_COL
                )

                cell.fill = retailer_fill
                cell.border = border

            retailer_cell.fill = retailer_fill
            retailer_cell.font = bold
            retailer_cell.alignment = center
            retailer_cell.border = border

            # D в строке заголовка
            ws.cell(
                row=header_row,
                column=START_COL + 1
            ).border = border

            # ==================================================
            # Названия показателей
            # ==================================================

            labels = [
                "DMR",
                f"OSA After {current_year}",
                f"OSA After {next_year}",
                "dif",
                "ЭФ-т LSV, %",
                "ЭФ-т, млн руб"
            ]

            for index, label in enumerate(labels):

                row = header_row + 1 + index

                cell = ws.cell(
                    row=row,
                    column=START_COL + 1,
                    value=label
                )

                cell.alignment = center
                cell.border = border

                if label == "ЭФ-т, млн руб":
                    cell.fill = green
                    cell.font = bold

            # ==================================================
            # DMR
            # ==================================================

            for i in range(13):

                period_name = f"p{i + 1}"

                dmr = dmr_values.get(
                    period_name
                )

                if (
                    dmr is not None
                    and isinstance(
                        dmr,
                        (int, float)
                    )
                    and not isinstance(
                        dmr,
                        bool
                    )
                ):

                    cell = ws.cell(
                        row=header_row + 1,
                        column=START_COL + 2 + i,
                        value=dmr
                    )

                    cell.number_format = "# ##0.00"
                    cell.border = border
                    cell.alignment = center

            # ==================================================
            # OSA After current_year
            # ==================================================

            osa_values = []

            for i, period in enumerate(periods):

                data = df[
                    (df["retailer_name"] == retailer)
                    &
                    (df["demand_region"] == region)
                    &
                    (df["full_period_name"] == period)
                ]

                osa = None

                if not data.empty:

                    value = data["OSA After"].iloc[0]

                    if pd.notna(value):

                        try:
                            osa = float(value)
                        except (
                            ValueError,
                            TypeError
                        ):
                            osa = None

                osa_values.append(osa)

                if osa is not None:

                    cell = ws.cell(
                        row=header_row + 2,
                        column=START_COL + 2 + i,
                        value=round(
                            osa,
                            4
                        )
                    )

                    cell.number_format = "0.00%"
                    cell.border = border
                    cell.alignment = center

            # ==================================================
            # Заполняем пропуски OSA текущего года Python
            # ==================================================

            osa_current_full = (
                forecast_missing_periods_osa(
                    osa_values
                )
            )

            for i in range(13):

                if (
                    osa_current_full[i]
                    is None
                ):
                    continue

                cell = ws.cell(
                    row=header_row + 2,
                    column=START_COL + 2 + i
                )

                # Если исходное значение уже есть,
                # не перезаписываем его.
                if cell.value is None:

                    cell.value = round(
                        float(
                            osa_current_full[i]
                        ),
                        4
                    )

                    cell.number_format = "0.00%"
                    cell.font = red_text
                    cell.border = border
                    cell.alignment = center

            # ==================================================
            # OSA After next_year
            #
            # ЭТО ОСТАЁТСЯ PYTHON-ЗНАЧЕНИЕМ
            # ==================================================

            full_osa_values = []

            for i in range(13):

                value = ws.cell(
                    row=header_row + 2,
                    column=START_COL + 2 + i
                ).value

                if value is not None:

                    full_osa_values.append(
                        float(value)
                    )

            osa_forecast = (
                forecast_next_year_osa(
                    full_osa_values
                )
            )

            for i in range(13):

                if i >= len(
                    osa_forecast
                ):
                    continue

                value = osa_forecast[i]

                if value is None:
                    continue

                cell = ws.cell(
                    row=header_row + 3,
                    column=START_COL + 2 + i,
                    value=round(
                        float(value),
                        4
                    )
                )

                cell.number_format = "0.00%"
                cell.font = red_text
                cell.border = border
                cell.alignment = center

            # ==================================================
            # EXCEL FORMULAS
            #
            # dif = OSA 2026 - OSA 2025
            # LSV = dif / 3% * 1%
            # ЭФ-т = DMR * LSV
            # ==================================================

            for i in range(13):

                col = START_COL + 2 + i

                column_letter = (
                    get_column_letter(col)
                )

                # ----------------------------------------------
                # DIF
                # ----------------------------------------------

                current_osa_cell = (
                    f"{column_letter}{header_row + 2}"
                )

                next_osa_cell = (
                    f"{column_letter}{header_row + 3}"
                )

                dif_cell = ws.cell(
                    row=header_row + 4,
                    column=col
                )

                dif_cell.value = (
                    f'=IF(OR('
                    f'{current_osa_cell}="",'
                    f'{next_osa_cell}=""),'
                    f'"",'
                    f'{next_osa_cell}-'
                    f'{current_osa_cell})'
                )

                dif_cell.number_format = "0.00%"
                dif_cell.border = border
                dif_cell.alignment = center

                # ----------------------------------------------
                # LSV
                # ----------------------------------------------

                lsv_cell = ws.cell(
                    row=header_row + 5,
                    column=col
                )

                lsv_cell.value = (
                    f'=IF({column_letter}{header_row + 4}="","",'
                    f'{column_letter}{header_row + 4}/3%*1%)'
                )

                lsv_cell.number_format = "0.00%"
                lsv_cell.border = border
                lsv_cell.alignment = center

                # ----------------------------------------------
                # ЭФ-т, млн руб
                # ----------------------------------------------

                effect_cell = ws.cell(
                    row=header_row + 6,
                    column=col
                )

                effect_cell.value = (
                    f'=IF(OR('
                    f'{column_letter}{header_row + 1}="",'
                    f'{column_letter}{header_row + 5}=""),'
                    f'"",'
                    f'{column_letter}{header_row + 1}*'
                    f'{column_letter}{header_row + 5})'
                )

                effect_cell.number_format = "# ##0.00"
                effect_cell.border = border
                effect_cell.alignment = center

            # ==================================================
            # TOTAL — тоже Excel formulas
            # ==================================================

            total_letter = get_column_letter(
                TOTAL_COL
            )

            # --------------------------------------------------
            # DMR total
            # --------------------------------------------------

            dmr_range = (
                f"{get_column_letter(START_COL + 2)}"
                f"{header_row + 1}:"
                f"{get_column_letter(START_COL + 14)}"
                f"{header_row + 1}"
            )

            dmr_total = ws.cell(
                row=header_row + 1,
                column=TOTAL_COL
            )

            dmr_total.value = (
                f'=IF(COUNT({dmr_range})=0,"",'
                f'SUM({dmr_range}))'
            )

            dmr_total.number_format = "# ##0.00"
            dmr_total.font = bold
            dmr_total.border = border
            dmr_total.alignment = center

            # --------------------------------------------------
            # OSA current total
            # --------------------------------------------------

            current_osa_range = (
                f"{get_column_letter(START_COL + 2)}"
                f"{header_row + 2}:"
                f"{get_column_letter(START_COL + 14)}"
                f"{header_row + 2}"
            )

            current_total = ws.cell(
                row=header_row + 2,
                column=TOTAL_COL
            )

            current_total.value = (
                f'=IF(COUNT({current_osa_range})=0,"",'
                f'AVERAGE({current_osa_range}))'
            )

            current_total.number_format = "0.00%"
            current_total.border = border
            current_total.alignment = center

            # --------------------------------------------------
            # OSA next total
            # --------------------------------------------------

            next_osa_range = (
                f"{get_column_letter(START_COL + 2)}"
                f"{header_row + 3}:"
                f"{get_column_letter(START_COL + 14)}"
                f"{header_row + 3}"
            )

            next_total = ws.cell(
                row=header_row + 3,
                column=TOTAL_COL
            )

            next_total.value = (
                f'=IF(COUNT({next_osa_range})=0,"",'
                f'AVERAGE({next_osa_range}))'
            )

            next_total.number_format = "0.00%"
            next_total.font = red_text
            next_total.border = border
            next_total.alignment = center

            # --------------------------------------------------
            # DIF total
            # --------------------------------------------------

            dif_total = ws.cell(
                row=header_row + 4,
                column=TOTAL_COL
            )

            dif_total.value = (
                f'=IF(OR('
                f'{total_letter}{header_row + 2}="",'
                f'{total_letter}{header_row + 3}=""),'
                f'"",'
                f'{total_letter}{header_row + 3}-'
                f'{total_letter}{header_row + 2})'
            )

            dif_total.number_format = "0.00%"
            dif_total.border = border
            dif_total.alignment = center

            # --------------------------------------------------
            # LSV total
            # --------------------------------------------------

            lsv_total = ws.cell(
                row=header_row + 5,
                column=TOTAL_COL
            )

            lsv_total.value = (
                f'=IF({total_letter}{header_row + 4}="","",'
                f'{total_letter}{header_row + 4}/3%*1%)'
            )

            lsv_total.number_format = "0.00%"
            lsv_total.border = border
            lsv_total.alignment = center

            # --------------------------------------------------
            # ЭФ-т total
            # --------------------------------------------------

            effect_range = (
                f"{get_column_letter(START_COL + 2)}"
                f"{header_row + 6}:"
                f"{get_column_letter(START_COL + 14)}"
                f"{header_row + 6}"
            )

            effect_total = ws.cell(
                row=header_row + 6,
                column=TOTAL_COL
            )

            effect_total.value = (
                f'=IF(COUNT({effect_range})=0,"",'
                f'SUM({effect_range}))'
            )

            effect_total.number_format = "# ##0.00"
            effect_total.border = border
            effect_total.alignment = center

            # ==================================================
            # Общие границы блока
            # ==================================================

            for row in range(
                header_row,
                header_row + 7
            ):

                for col in range(
                    START_COL,
                    TOTAL_COL + 1
                ):

                    cell = ws.cell(
                        row=row,
                        column=col
                    )

                    cell.border = border

                    if cell.alignment == Alignment():
                        cell.alignment = center

            # ==================================================
            # Зелёное выделение строки ЭФ-т
            # ==================================================

            effect_row = header_row + 6

            for col in range(
                START_COL,
                TOTAL_COL + 1
            ):

                cell = ws.cell(
                    row=effect_row,
                    column=col
                )

                cell.fill = green
                cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=medium
                )

                if col == START_COL + 1:
                    cell.font = bold

            # ==================================================
            # Красные аналитические расчёты
            # ==================================================

            for row in [
                header_row + 3,  # OSA next year
                header_row + 4,  # dif
                header_row + 5   # LSV
            ]:

                for col in range(
                    START_COL + 2,
                    TOTAL_COL + 1
                ):

                    cell = ws.cell(
                        row=row,
                        column=col
                    )


            # ==================================================
            # Следующая компания
            # ==================================================

            current_row += 7

        # ======================================================
        # Следующий регион
        # ======================================================

        current_row += 1

    # ==========================================================
    # Убираем пустые строки внизу
    # ==========================================================

    max_row = ws.max_row

    while max_row > START_ROW:

        values = [
            ws.cell(
                row=max_row,
                column=col
            ).value
            for col in range(
                START_COL,
                TOTAL_COL + 1
            )
        ]

        if any(
            value is not None
            for value in values
        ):
            break

        max_row -= 1

    # ==========================================================
    # Настройки листа
    # ==========================================================

    ws.freeze_panes = "E4"

    return wb
